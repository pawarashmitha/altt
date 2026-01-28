from flask import Flask, request, jsonify, send_file, render_template_string, session, redirect, url_for
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this-in-production'  # Change this!
CORS(app, supports_credentials=True)

EXCEL_FILE = 'GP_Metrics_Historic.xlsx'

# GP Login Credentials (In production, use a database and hashed passwords)
GP_CREDENTIALS = {
    'BlackRock': 'blackrock123',
    'Brookfield': 'brookfield123',
    'TPG': 'tpg123'
}

LOGIN_HTML = '''
<!DOCTYPE html>
<html>
<head>
    <title>GP Metrics Tracker - Login</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 20px;
        }
        .login-container {
            background: white;
            padding: 3rem;
            border-radius: 16px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            width: 100%;
            max-width: 400px;
        }
        h1 {
            color: #1f2937;
            margin-bottom: 0.5rem;
            text-align: center;
        }
        .subtitle {
            color: #6b7280;
            text-align: center;
            margin-bottom: 2rem;
            font-size: 0.875rem;
        }
        .form-group {
            margin-bottom: 1.5rem;
        }
        label {
            display: block;
            color: #374151;
            font-weight: 500;
            margin-bottom: 0.5rem;
            font-size: 0.875rem;
        }
        select, input {
            width: 100%;
            padding: 0.75rem;
            border: 2px solid #e5e7eb;
            border-radius: 8px;
            font-size: 1rem;
            transition: border-color 0.2s;
        }
        select:focus, input:focus {
            outline: none;
            border-color: #667eea;
        }
        .btn {
            width: 100%;
            padding: 0.875rem;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            border-radius: 8px;
            font-size: 1rem;
            font-weight: 600;
            cursor: pointer;
            transition: transform 0.2s;
        }
        .btn:hover {
            transform: translateY(-2px);
        }
        .alert {
            padding: 0.75rem;
            border-radius: 8px;
            margin-bottom: 1rem;
            display: none;
        }
        .alert.show { display: block; }
        .alert-error {
            background: #fee2e2;
            color: #991b1b;
            border-left: 4px solid #ef4444;
        }
        .credentials {
            background: #f9fafb;
            padding: 1rem;
            border-radius: 8px;
            margin-top: 2rem;
            font-size: 0.75rem;
        }
        .credentials h3 {
            color: #374151;
            font-size: 0.875rem;
            margin-bottom: 0.5rem;
        }
        .credentials p {
            color: #6b7280;
            margin: 0.25rem 0;
        }
        .lock-icon {
            text-align: center;
            font-size: 3rem;
            margin-bottom: 1rem;
        }
    </style>
</head>
<body>
    <div class="login-container">
        <div class="lock-icon">🔒</div>
        <h1>GP Metrics Tracker</h1>
        <p class="subtitle">Secure Access Portal</p>
        
        <div id="alert" class="alert"></div>
        
        <form id="loginForm">
            <div class="form-group">
                <label>Select Your GP</label>
                <select id="gpName" required>
                    <option value="">-- Select GP --</option>
                    <option value="BlackRock">BlackRock</option>
                    <option value="Brookfield">Brookfield</option>
                    <option value="TPG">TPG</option>
                </select>
            </div>
            
            <div class="form-group">
                <label>Password</label>
                <input type="password" id="password" placeholder="Enter your password" required>
            </div>
            
            <button type="submit" class="btn">Login</button>
        </form>
        
        <div class="credentials">
            <h3>Demo Credentials:</h3>
            <p><strong>BlackRock:</strong> blackrock123</p>
            <p><strong>Brookfield:</strong> brookfield123</p>
            <p><strong>TPG:</strong> tpg123</p>
        </div>
    </div>

    <script>
        document.getElementById('loginForm').addEventListener('submit', async (e) => {
            e.preventDefault();
            
            const gpName = document.getElementById('gpName').value;
            const password = document.getElementById('password').value;
            
            if (!gpName) {
                showAlert('Please select a GP');
                return;
            }
            
            try {
                const response = await fetch('/api/login', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ gpName, password })
                });
                
                const result = await response.json();
                
                if (result.success) {
                    window.location.href = '/frontend';
                } else {
                    showAlert(result.error || 'Invalid credentials');
                }
            } catch (error) {
                showAlert('Connection error');
            }
        });
        
        function showAlert(message) {
            const alert = document.getElementById('alert');
            alert.textContent = message;
            alert.className = 'alert alert-error show';
            setTimeout(() => alert.className = 'alert', 5000);
        }
    </script>
</body>
</html>
'''

FRONTEND_HTML = '''
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>GP Metrics Tracker - {{ gp_name }}</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: linear-gradient(135deg, #e0f2fe 0%, #ddd6fe 100%);
            padding: 2rem;
        }
        .container { max-width: 1200px; margin: 0 auto; }
        .card {
            background: white;
            border-radius: 12px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
            padding: 2rem;
            margin-bottom: 2rem;
        }
        .header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 2rem;
            flex-wrap: wrap;
            gap: 1rem;
        }
        h1 { font-size: 2rem; color: #1f2937; }
        .gp-badge {
            display: inline-block;
            background: #dbeafe;
            color: #1e40af;
            padding: 0.5rem 1rem;
            border-radius: 20px;
            font-weight: 600;
            font-size: 0.875rem;
        }
        .subtitle { color: #6b7280; margin-top: 0.5rem; }
        .section {
            margin-bottom: 2rem;
            padding-bottom: 2rem;
            border-bottom: 1px solid #e5e7eb;
        }
        .section:last-child { border-bottom: none; }
        .section-title {
            font-size: 1.25rem;
            font-weight: 600;
            color: #374151;
            margin-bottom: 1rem;
        }
        .form-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 1rem;
        }
        .form-field { display: flex; flex-direction: column; }
        .form-field.full-width { grid-column: 1 / -1; }
        label {
            font-size: 0.875rem;
            font-weight: 500;
            color: #374151;
            margin-bottom: 0.5rem;
        }
        .required { color: #ef4444; }
        input, select, textarea {
            padding: 0.75rem;
            border: 1px solid #d1d5db;
            border-radius: 8px;
            font-size: 1rem;
        }
        input:focus, select:focus, textarea:focus {
            outline: none;
            border-color: #2563eb;
            box-shadow: 0 0 0 3px rgba(37, 99, 235, 0.1);
        }
        textarea { resize: vertical; min-height: 80px; }
        .btn {
            padding: 0.75rem 1.5rem;
            border: none;
            border-radius: 8px;
            font-size: 1rem;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.2s;
        }
        .btn-primary {
            background: #2563eb;
            color: white;
            width: 100%;
            padding: 1rem;
        }
        .btn-primary:hover { background: #1d4ed8; }
        .btn-success { background: #16a34a; color: white; }
        .btn-success:hover { background: #15803d; }
        .btn-logout {
            background: #ef4444;
            color: white;
        }
        .btn-logout:hover { background: #dc2626; }
        .alert {
            padding: 1rem;
            border-radius: 8px;
            margin-bottom: 1.5rem;
            display: none;
        }
        .alert.show { display: block; }
        .alert-success { background: #dcfce7; color: #166534; }
        .alert-error { background: #fee2e2; color: #991b1b; }
        .top-bar {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 1rem;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            font-size: 0.875rem;
        }
        th, td {
            padding: 0.75rem;
            text-align: left;
            border-bottom: 1px solid #e5e7eb;
        }
        th {
            background: #f9fafb;
            font-weight: 600;
            color: #374151;
        }
        tr:hover { background: #f9fafb; }
    </style>
</head>
<body>
    <div class="container">
        <div class="card">
            <div class="top-bar">
                <div>
                    <span class="gp-badge">{{ gp_name }}</span>
                </div>
                <button class="btn btn-logout" onclick="logout()">Logout</button>
            </div>
            
            <div class="header">
                <div>
                    <h1>Quarterly Metrics Tracker</h1>
                    <p class="subtitle">Energy & Sustainability Performance Data</p>
                </div>
                <button class="btn btn-success" onclick="downloadExcel()">📊 Download Your Data</button>
            </div>

            <div id="alert" class="alert"></div>

            <div class="form-container">
                <div class="section">
                    <h2 class="section-title">Project Information</h2>
                    <div class="form-grid">
                        <div class="form-field">
                            <label>Project Name <span class="required">*</span></label>
                            <input type="text" id="projectName" required>
                        </div>
                        <div class="form-field">
                            <label>Date <span class="required">*</span></label>
                            <input type="date" id="date" required>
                        </div>
                        <div class="form-field">
                            <label>Tech Type <span class="required">*</span></label>
                            <input type="text" id="techType" required>
                        </div>
                        <div class="form-field">
                            <label>Asset <span class="required">*</span></label>
                            <input type="text" id="asset" required>
                        </div>
                        <div class="form-field">
                            <label>Location <span class="required">*</span></label>
                            <input type="text" id="location" required>
                        </div>
                    </div>
                </div>

                <div class="section">
                    <h2 class="section-title">Energy Capacity & Generation</h2>
                    <div class="form-grid">
                        <div class="form-field">
                            <label>Planned Capacity (MW)</label>
                            <input type="number" id="plannedCapacityMW" step="0.01">
                        </div>
                        <div class="form-field">
                            <label>Actual Capacity (MW)</label>
                            <input type="number" id="actualCapacityMW" step="0.01">
                        </div>
                        <div class="form-field">
                            <label>Planned Energy (MWh)</label>
                            <input type="number" id="plannedEnergyMWh" step="0.01">
                        </div>
                        <div class="form-field">
                            <label>Actual Energy (MWh)</label>
                            <input type="number" id="actualEnergyMWh" step="0.01">
                        </div>
                    </div>
                </div>

                <div class="section">
                    <h2 class="section-title">Performance Metrics</h2>
                    <div class="form-grid">
                        <div class="form-field">
                            <label>Capacity Factor (%)</label>
                            <input type="number" id="capacityFactor" step="0.01">
                        </div>
                        <div class="form-field">
                            <label>Availability (%)</label>
                            <input type="number" id="availability" step="0.01">
                        </div>
                        <div class="form-field">
                            <label>Generation Efficiency (%)</label>
                            <input type="number" id="generationEfficiency" step="0.01">
                        </div>
                        <div class="form-field">
                            <label>Unavailability Rate (%)</label>
                            <input type="number" id="unavailabilityRate" step="0.01">
                        </div>
                    </div>
                </div>
                
                <button class="btn btn-primary" onclick="submitForm()">Submit Metrics</button>
            </div>
        </div>

        <div class="card" id="recordsCard" style="display: none;">
            <h2 class="section-title">Your Submitted Records</h2>
            <div id="recordsTable"></div>
        </div>
    </div>

    <script>
        const gpName = "{{ gp_name }}";

        window.onload = function() {
            loadRecords();
        };

        async function submitForm() {
            const data = {
                gpName: gpName,
                projectName: document.getElementById('projectName').value,
                date: document.getElementById('date').value,
                techType: document.getElementById('techType').value,
                asset: document.getElementById('asset').value,
                location: document.getElementById('location').value,
                plannedCapacityMW: document.getElementById('plannedCapacityMW').value,
                actualCapacityMW: document.getElementById('actualCapacityMW').value,
                plannedEnergyMWh: document.getElementById('plannedEnergyMWh').value,
                actualEnergyMWh: document.getElementById('actualEnergyMWh').value,
                capacityFactor: document.getElementById('capacityFactor').value,
                availability: document.getElementById('availability').value,
                generationEfficiency: document.getElementById('generationEfficiency').value,
                unavailabilityRate: document.getElementById('unavailabilityRate').value
            };

            if (!data.projectName || !data.date || !data.techType || !data.asset || !data.location) {
                showAlert('Please fill in all required fields', 'error');
                return;
            }

            try {
                const response = await fetch('/api/submit', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify(data)
                });

                const result = await response.json();
                if (result.success) {
                    showAlert('Metrics submitted successfully!', 'success');
                    document.querySelectorAll('input[type="text"], input[type="number"], input[type="date"]').forEach(el => {
                        el.value = '';
                    });
                    loadRecords();
                } else {
                    showAlert('Submission failed', 'error');
                }
            } catch (error) {
                showAlert('Connection error', 'error');
            }
        }

        async function loadRecords() {
            try {
                const response = await fetch('/api/records');
                const records = await response.json();
                
                if (records && records.length > 0) {
                    document.getElementById('recordsCard').style.display = 'block';
                    let html = '<table><thead><tr>';
                    html += '<th>Date</th><th>Project</th><th>Location</th><th>Tech Type</th><th>Capacity (MW)</th>';
                    html += '</tr></thead><tbody>';
                    
                    records.forEach(record => {
                        html += '<tr>';
                        html += '<td>' + record.date + '</td>';
                        html += '<td>' + record.projectName + '</td>';
                        html += '<td>' + record.location + '</td>';
                        html += '<td>' + record.techType + '</td>';
                        html += '<td>' + record.actualCapacityMW + '</td>';
                        html += '</tr>';
                    });
                    
                    html += '</tbody></table>';
                    document.getElementById('recordsTable').innerHTML = html;
                }
            } catch (error) {
                console.error('Error loading records:', error);
            }
        }

        function showAlert(message, type) {
            const alert = document.getElementById('alert');
            alert.textContent = message;
            alert.className = 'alert alert-' + type + ' show';
            setTimeout(() => alert.className = 'alert', 5000);
        }

        function downloadExcel() {
            window.location.href = '/api/download';
        }

        function logout() {
            fetch('/api/logout', { method: 'POST' })
                .then(() => window.location.href = '/login');
        }
    </script>
</body>
</html>
'''

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        wb.remove(wb.active)
        
        for gp in ['BlackRock', 'Brookfield', 'TPG']:
            ws = wb.create_sheet(title=gp)
            headers = [
                'GP Name', 'Submitted At', 'Project Name', 'Date', 'Tech Type', 'Asset', 'Location',
                'Planned Capacity (MW)', 'Actual Capacity (MW)', 'Planned Energy (MWh)', 'Actual Energy (MWh)',
                'Capacity Factor (%)', 'Availability (%)', 'Generation Efficiency (%)', 'Unavailability Rate (%)',
                'Scheduled Downtime (hrs)', 'Unscheduled Downtime (hrs)', 'Downtime Reasons',
                'Planned Maintenance', 'Unexpected Maintenance', 'Compensated Curtailment (MWh)',
                'Uncompensated Curtailment (MWh)', 'Weather Impact', 'Performance Variance',
                'Technology', 'Total Extent (acres)', 'Degradation Rate (%)', 'Storage Utilization (MWh)',
                'Scope 1 Emissions (CO2e)', 'Scope 2 Emissions (CO2e)', 'Scope 3 Emissions (CO2e)',
                'Scope 3 Categories', 'Emissions Comments', 'Emission Targets', 'Target Status',
                'Target Deviation', 'Water Consumption (L)', 'Water Conservation', 'Biodiversity Assessment',
                'Jobs Created', 'Severe Injury Rate', 'Comments'
            ]
            ws.append(headers)
        
        wb.save(EXCEL_FILE)

def require_login(f):
    def wrapper(*args, **kwargs):
        if 'gp_name' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    wrapper.__name__ = f.__name__
    return wrapper

@app.route('/')
def home():
    return redirect(url_for('login'))

@app.route('/login')
def login():
    if 'gp_name' in session:
        return redirect(url_for('frontend'))
    return render_template_string(LOGIN_HTML)

@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.json
    gp_name = data.get('gpName')
    password = data.get('password')
    
    if gp_name in GP_CREDENTIALS and GP_CREDENTIALS[gp_name] == password:
        session['gp_name'] = gp_name
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'error': 'Invalid credentials'}), 401

@app.route('/api/logout', methods=['POST'])
def api_logout():
    session.pop('gp_name', None)
    return jsonify({'success': True})

@app.route('/frontend')
@require_login
def frontend():
    gp_name = session.get('gp_name')
    return render_template_string(FRONTEND_HTML, gp_name=gp_name)

@app.route('/api/submit', methods=['POST'])
@require_login
def submit_metrics():
    try:
        data = request.json
        logged_in_gp = session.get('gp_name')
        
        # Security: Only allow GP to submit their own data
        if data.get('gpName') != logged_in_gp:
            return jsonify({'success': False, 'error': 'Unauthorized'}), 403
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb[logged_in_gp]
        
        row = [
            logged_in_gp,
            datetime.now().isoformat(),
            data.get('projectName', ''),
            data.get('date', ''),
            data.get('techType', ''),
            data.get('asset', ''),
            data.get('location', ''),
            data.get('plannedCapacityMW', ''),
            data.get('actualCapacityMW', ''),
            data.get('plannedEnergyMWh', ''),
            data.get('actualEnergyMWh', ''),
            data.get('capacityFactor', ''),
            data.get('availability', ''),
            data.get('generationEfficiency', ''),
            data.get('unavailabilityRate', ''),
            data.get('scheduledDowntime', ''),
            data.get('unscheduledDowntime', ''),
            data.get('downtimeReasons', ''),
            data.get('plannedMaintenance', ''),
            data.get('unexpectedMaintenance', ''),
            data.get('compensatedCurtailment', ''),
            data.get('uncompensatedCurtailment', ''),
            data.get('weatherImpact', ''),
            data.get('performanceVariance', ''),
            data.get('technology', ''),
            data.get('totalExtentAcres', ''),
            data.get('degradationRate', ''),
            data.get('storageUtilizationMWh', ''),
            data.get('scope1Emissions', ''),
            data.get('scope2Emissions', ''),
            data.get('scope3Emissions', ''),
            data.get('scope3Categories', ''),
            data.get('emissionsComments', ''),
            data.get('emissionTargets', ''),
            data.get('targetStatus', ''),
            data.get('targetDeviation', ''),
            data.get('waterConsumption', ''),
            data.get('waterConservation', ''),
            data.get('biodiversityAssessment', ''),
            data.get('jobsCreated', ''),
            data.get('severeInjuryRate', ''),
            data.get('comments', '')
        ]
        
        ws.append(row)
        wb.save(EXCEL_FILE)
        
        return jsonify({'success': True, 'message': 'Data added successfully'})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/download', methods=['GET'])
@require_login
def download_excel():
    try:
        logged_in_gp = session.get('gp_name')
        
        # Create a new workbook with only the logged-in GP's data
        wb_full = load_workbook(EXCEL_FILE)
        wb_filtered = Workbook()
        wb_filtered.remove(wb_filtered.active)
        
        # Copy only the logged-in GP's sheet
        ws_source = wb_full[logged_in_gp]
        ws_dest = wb_filtered.create_sheet(title=logged_in_gp)
        
        for row in ws_source.iter_rows(values_only=True):
            ws_dest.append(row)
        
        # Save to temporary file
        temp_file = f'{logged_in_gp}_Metrics_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb_filtered.save(temp_file)
        
        return send_file(temp_file, as_attachment=True, download_name=temp_file)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/records', methods=['GET'])
@require_login
def get_records():
    try:
        logged_in_gp = session.get('gp_name')
        wb = load_workbook(EXCEL_FILE)
        ws = wb[logged_in_gp]
        
        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                records.append({
                    'gpName': row[0],
                    'submittedAt': row[1],
                    'projectName': row[2],
                    'date': row[3],
                    'techType': row[4],
                    'location': row[6],
                    'actualCapacityMW': row[8]
                })
        
        return jsonify(records)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    init_excel()
    print("=" * 60)
    print("🔒 GP Metrics Tracker with Authentication Started!")
    print("=" * 60)
    print("📍 Open in browser: http://localhost:5000")
    print("")
    print("🔑 Login Credentials:")
    print("   BlackRock: blackrock123")
    print("   Brookfield: brookfield123")
    print("   TPG: tpg123")
    print("=" * 60)
    app.run(debug=True, port=5000)